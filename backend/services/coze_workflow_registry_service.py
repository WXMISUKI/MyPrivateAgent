"""Read-only registry for Coze migration workflow manifests."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional
from uuid import uuid4

CONTRACT_VERSION = "coze-workflow-registry-v1"
DEFAULT_COZE_WORKFLOW_ROOT = Path(__file__).resolve().parents[1] / "coze_workflows"
SUPPORTED_STATUSES = {"draft", "review", "active", "deprecated", "archived"}
SUPPORTED_RUNTIME_CAPABILITIES = {
    "document.file_type.detect",
    "http.request",
    "spreadsheet.table.extract",
    "llm.structured_json.generate",
    "json_schema.validate",
}


@dataclass(frozen=True)
class _ManifestResult:
    workflow: Dict[str, Any] | None
    error: Dict[str, Any] | None


class CozeWorkflowRegistryService:
    """Discover Coze migration workflow manifests without importing or executing workflow code."""

    def __init__(self, root_path: Path | str | None = None):
        self.root_path = Path(root_path) if root_path is not None else DEFAULT_COZE_WORKFLOW_ROOT

    def build_runtime_contract(self) -> Dict[str, Any]:
        workflows: List[Dict[str, Any]] = []
        errors: List[Dict[str, Any]] = []

        for manifest_path in self._iter_manifest_paths():
            result = self._load_manifest(manifest_path)
            if result.workflow is not None:
                workflows.append(result.workflow)
            if result.error is not None:
                errors.append(result.error)

        ready_workflows = sum(1 for w in workflows if w.get("readiness", {}).get("status") == "ready")
        invalid_workflows = len(errors) + sum(1 for w in workflows if w.get("readiness", {}).get("status") == "invalid")
        total_workflows = len(workflows) + len(errors)

        if total_workflows == 0:
            status = "empty"
        elif invalid_workflows:
            status = "degraded"
        else:
            status = "ready"

        workflows.sort(key=lambda item: str(item.get("id") or ""))
        errors.sort(key=lambda item: str(item.get("manifest_path") or ""))
        capabilities = [
            self._build_capability_contract(workflow)
            for workflow in workflows
            if workflow.get("readiness", {}).get("status") == "ready"
        ]

        return {
            "contract_version": CONTRACT_VERSION,
            "status": status,
            "root_path": str(self.root_path),
            "total_workflows": total_workflows,
            "ready_workflows": ready_workflows,
            "invalid_workflows": invalid_workflows,
            "workflows": workflows,
            "capabilities": capabilities,
            "errors": errors,
        }

    def get_workflow_by_id(self, workflow_id: str) -> Dict[str, Any] | None:
        contract = self.build_runtime_contract()
        for workflow in contract.get("workflows", []):
            if workflow.get("id") == workflow_id:
                return workflow
        return None

    def build_capability_definitions(self) -> List["CapabilityDefinition"]:
        from backend.capability_runtime.contracts import CapabilityDefinition

        definitions: List[CapabilityDefinition] = []
        for workflow in self.build_runtime_contract().get("workflows", []):
            if workflow.get("readiness", {}).get("status") != "ready":
                continue
            workflow_id = str(workflow.get("id") or "").strip()
            if not workflow_id:
                continue
            definitions.append(
                CapabilityDefinition(
                    capability_id=str(workflow.get("capability_id") or f"coze.workflow.{workflow_id}"),
                    kind="workflow",
                    transport="local",
                    provider="coze_migration",
                    title=str(workflow.get("name") or workflow_id),
                    description=self._build_capability_description(workflow),
                    input_schema=dict(workflow.get("inputs", {}).get("schema") or {}),
                    output_schema=dict(workflow.get("outputs", {}).get("schema") or {}),
                    metadata=self._build_capability_metadata(workflow),
                    invoker=lambda payload, workflow_id=workflow_id: self.invoke_workflow(workflow_id, payload),
                    health_checker=lambda workflow=workflow: {
                        "status": str(workflow.get("readiness", {}).get("status") or "unknown"),
                        "reason": str(workflow.get("readiness", {}).get("reason") or ""),
                    },
                )
            )
        return definitions

    def invoke_workflow(self, workflow_id: str, payload: Mapping[str, Any] | None) -> Dict[str, Any]:
        workflow = self.get_workflow_by_id(workflow_id)
        if workflow is None:
            return self._build_error_envelope(
                workflow_id=workflow_id,
                capability_id=f"coze.workflow.{workflow_id}",
                workflow_version="",
                status="not_found",
                code="COZE_WORKFLOW_NOT_FOUND",
                message=f"Coze workflow not found: {workflow_id}",
            )

        workflow_version = str(workflow.get("version") or "")
        readiness = dict(workflow.get("readiness") or {})
        if readiness.get("status") == "invalid" or workflow.get("status") == "invalid":
            return self._build_error_envelope(
                workflow_id=workflow_id,
                capability_id=str(workflow.get("capability_id") or f"coze.workflow.{workflow_id}"),
                workflow_version=workflow_version,
                status="invalid",
                code="COZE_WORKFLOW_INVALID_MANIFEST",
                message="Workflow manifest is invalid.",
                blockers=list(readiness.get("blockers") or []),
                details={"reason": readiness.get("reason")},
            )

        if readiness.get("status") != "ready":
            code = "COZE_WORKFLOW_BLOCKED"
            if any(str(item).startswith("missing_") for item in (readiness.get("blockers") or [])):
                code = "COZE_WORKFLOW_DEPENDENCY_UNAVAILABLE"
            return self._build_error_envelope(
                workflow_id=workflow_id,
                capability_id=str(workflow.get("capability_id") or f"coze.workflow.{workflow_id}"),
                workflow_version=workflow_version,
                status=str(readiness.get("status") or "blocked"),
                code=code,
                message="Workflow is not ready for invocation.",
                blockers=list(readiness.get("blockers") or []),
                details={"reason": readiness.get("reason")},
            )

        normalized_payload = dict(payload or {})
        schema_errors = self._validate_payload_against_schema(
            dict(workflow.get("inputs", {}).get("schema") or {}),
            normalized_payload,
        )
        if schema_errors:
            return self._build_error_envelope(
                workflow_id=workflow_id,
                capability_id=str(workflow.get("capability_id") or f"coze.workflow.{workflow_id}"),
                workflow_version=workflow_version,
                status="invalid_input",
                code="COZE_WORKFLOW_SCHEMA_VALIDATION_FAILED",
                message="Workflow input payload failed schema validation.",
                details={"validation_errors": schema_errors},
            )

        dependency_blockers = self._detect_dependency_blockers(workflow)
        if dependency_blockers:
            return self._build_error_envelope(
                workflow_id=workflow_id,
                capability_id=str(workflow.get("capability_id") or f"coze.workflow.{workflow_id}"),
                workflow_version=workflow_version,
                status="blocked",
                code="COZE_WORKFLOW_DEPENDENCY_UNAVAILABLE",
                message="Workflow dependencies are unavailable.",
                blockers=dependency_blockers,
                details={"dependency_summary": self._build_dependency_summary(workflow)},
            )

        try:
            result = self._execute_workflow(workflow, normalized_payload)
        except FileNotFoundError as exc:
            return self._build_error_envelope(
                workflow_id=workflow_id,
                capability_id=str(workflow.get("capability_id") or f"coze.workflow.{workflow_id}"),
                workflow_version=workflow_version,
                status="blocked",
                code="COZE_WORKFLOW_DEPENDENCY_UNAVAILABLE",
                message=str(exc),
                details={"dependency_summary": self._build_dependency_summary(workflow)},
            )
        except RuntimeError as exc:
            return self._build_error_envelope(
                workflow_id=workflow_id,
                capability_id=str(workflow.get("capability_id") or f"coze.workflow.{workflow_id}"),
                workflow_version=workflow_version,
                status="blocked",
                code="COZE_WORKFLOW_EXECUTOR_UNAVAILABLE",
                message=str(exc),
                details={"dependency_summary": self._build_dependency_summary(workflow)},
            )
        except ValueError as exc:
            return self._build_error_envelope(
                workflow_id=workflow_id,
                capability_id=str(workflow.get("capability_id") or f"coze.workflow.{workflow_id}"),
                workflow_version=workflow_version,
                status="invalid_input",
                code="COZE_WORKFLOW_EXECUTION_FAILED",
                message=str(exc),
                details={"dependency_summary": self._build_dependency_summary(workflow)},
            )

        run_id = f"run_{uuid4().hex}"
        dependency_summary = self._build_dependency_summary(workflow)
        trace_summary = {
            "workflow_id": workflow_id,
            "workflow_version": workflow_version,
            "owner": str(workflow.get("owner", {}).get("primary") or ""),
            "source": "coze_migration",
            "dependency_summary": dependency_summary,
            "manifest_path": workflow.get("manifest_path"),
            "workflow_dir": workflow.get("workflow_dir"),
        }
        return {
            "ok": True,
            "workflow_id": workflow_id,
            "capability_id": str(workflow.get("capability_id") or f"coze.workflow.{workflow_id}"),
            "workflow_version": workflow_version,
            "run_id": run_id,
            "status": "completed",
            "result": result,
            "authorization": self._build_authorization_placeholder(workflow),
            "invocation_policy": self._build_invocation_policy(workflow),
            "trace_summary": trace_summary,
            "trace": trace_summary,
        }

    def _build_capability_contract(self, workflow: Mapping[str, Any]) -> Dict[str, Any]:
        readiness = dict(workflow.get("readiness") or {})
        return {
            "capability_id": str(workflow.get("capability_id") or ""),
            "kind": "workflow",
            "workflow_id": str(workflow.get("id") or ""),
            "title": str(workflow.get("name") or ""),
            "status": str(readiness.get("status") or "unknown"),
            "reason": str(readiness.get("reason") or ""),
            "input_schema": dict(workflow.get("inputs", {}).get("schema") or {}),
            "output_schema": dict(workflow.get("outputs", {}).get("schema") or {}),
            "owner": workflow.get("owner", {}).get("primary"),
            "version": str(workflow.get("version") or ""),
            "workflow_status": str(workflow.get("status") or ""),
            "governance": dict(workflow.get("governance") or {}),
            "readiness": {
                "status": str(readiness.get("status") or "unknown"),
                "reason": str(readiness.get("reason") or ""),
                "blockers": list(readiness.get("blockers") or []),
            },
            "asset_paths": {
                "workflow_dir": workflow.get("workflow_dir"),
                "manifest_path": workflow.get("manifest_path"),
            },
            "invocation_policy": self._build_invocation_policy(workflow),
        }

    def _build_capability_description(self, workflow: Mapping[str, Any]) -> str:
        source = workflow.get("source") if isinstance(workflow, Mapping) else {}
        migration_notes = str((source or {}).get("migration_notes") or "").strip()
        if migration_notes:
            return migration_notes
        return f"Coze migration workflow: {workflow.get('id')}"

    def _build_capability_metadata(self, workflow: Mapping[str, Any]) -> Dict[str, Any]:
        return {
            "workflow_id": str(workflow.get("id") or ""),
            "workflow_version": str(workflow.get("version") or ""),
            "workflow_status": str(workflow.get("status") or ""),
            "owner": dict(workflow.get("owner") or {}),
            "source": dict(workflow.get("source") or {}),
            "governance": dict(workflow.get("governance") or {}),
            "acceptance": dict(workflow.get("acceptance") or {}),
            "workflow_dir": workflow.get("workflow_dir"),
            "manifest_path": workflow.get("manifest_path"),
            "asset_paths": {
                "workflow_dir": workflow.get("workflow_dir"),
                "manifest_path": workflow.get("manifest_path"),
            },
            "readiness": dict(workflow.get("readiness") or {}),
            "invocation_policy": self._build_invocation_policy(workflow),
        }

    def _build_error_envelope(
        self,
        *,
        workflow_id: str,
        capability_id: str,
        workflow_version: str,
        status: str,
        code: str,
        message: str,
        blockers: List[str] | None = None,
        details: Mapping[str, Any] | None = None,
    ) -> Dict[str, Any]:
        return {
            "ok": False,
            "workflow_id": workflow_id,
            "capability_id": capability_id,
            "workflow_version": workflow_version,
            "run_id": None,
            "status": status,
            "error": {
                "code": code,
                "message": message,
                "blockers": list(blockers or []),
                "details": dict(details or {}),
            },
            "authorization": {
                "status": "not_evaluated",
                "policy": "placeholder",
                "reason": "Workflow API authorization is not implemented in this change.",
            },
            "invocation_policy": {
                "allowed_callers": [],
                "approval_required": False,
                "permission_level": "",
                "placeholder": True,
            },
            "trace_summary": {
                "workflow_id": workflow_id,
                "workflow_version": workflow_version,
                "source": "coze_migration",
            },
            "trace": {
                "workflow_id": workflow_id,
                "workflow_version": workflow_version,
                "source": "coze_migration",
            },
        }

    @staticmethod
    def _build_authorization_placeholder(workflow: Mapping[str, Any]) -> Dict[str, Any]:
        governance = dict(workflow.get("governance") or {})
        return {
            "status": "not_evaluated",
            "policy": "placeholder",
            "allowed_callers": list(governance.get("allowed_callers") or []),
            "reason": "Workflow API authorization is not implemented in this change.",
        }

    @staticmethod
    def _build_invocation_policy(workflow: Mapping[str, Any]) -> Dict[str, Any]:
        governance = dict(workflow.get("governance") or {})
        return {
            "allowed_callers": list(governance.get("allowed_callers") or []),
            "approval_required": bool(governance.get("approval_required")),
            "permission_level": str(governance.get("permission_level") or ""),
            "data_sensitivity": str(governance.get("data_sensitivity") or ""),
            "placeholder": True,
            "auth_system": "not_implemented_in_this_change",
        }

    def _build_dependency_summary(self, workflow: Mapping[str, Any]) -> Dict[str, Any]:
        dependencies = dict(workflow.get("dependencies") or {})
        return {
            "tools": list(dependencies.get("tools") or []),
            "mcp_capabilities": list(dependencies.get("mcp_capabilities") or []),
            "skills": list(dependencies.get("skills") or []),
            "providers": list(dependencies.get("providers") or []),
            "knowledge_sources": list(dependencies.get("knowledge_sources") or []),
            "runtime_capabilities": list(dependencies.get("runtime_capabilities") or []),
        }

    def _detect_dependency_blockers(self, workflow: Mapping[str, Any]) -> List[str]:
        blockers: List[str] = []
        dependencies = dict(workflow.get("dependencies") or {})
        runtime_capabilities = [
            str(item or "").strip()
            for item in (dependencies.get("runtime_capabilities") or [])
            if str(item or "").strip()
        ]
        missing_runtime_capabilities = sorted(
            cap for cap in runtime_capabilities if cap not in SUPPORTED_RUNTIME_CAPABILITIES
        )
        if missing_runtime_capabilities:
            blockers.append(
                "missing_runtime_capabilities:" + ",".join(missing_runtime_capabilities)
            )
        return blockers

    def _validate_payload_against_schema(self, schema: Mapping[str, Any], payload: Any, path: str = "payload") -> List[Dict[str, Any]]:
        if not schema:
            return []
        schema_type = schema.get("type")
        errors: List[Dict[str, Any]] = []
        if schema_type == "object":
            if not isinstance(payload, Mapping):
                return [{"path": path, "message": "Expected object"}]
            required = [str(item) for item in (schema.get("required") or []) if str(item).strip()]
            for key in required:
                if key not in payload:
                    errors.append({"path": f"{path}.{key}", "message": "Missing required field"})
            properties = dict(schema.get("properties") or {})
            for key, sub_schema in properties.items():
                if key in payload:
                    errors.extend(
                        self._validate_payload_against_schema(
                            sub_schema if isinstance(sub_schema, Mapping) else {},
                            payload.get(key),
                            path=f"{path}.{key}",
                        )
                    )
            return errors
        if schema_type == "array":
            if not isinstance(payload, list):
                return [{"path": path, "message": "Expected array"}]
            item_schema = schema.get("items") if isinstance(schema.get("items"), Mapping) else {}
            for index, item in enumerate(payload):
                errors.extend(self._validate_payload_against_schema(item_schema, item, path=f"{path}[{index}]"))
            return errors
        if schema_type == "string":
            if not isinstance(payload, str):
                return [{"path": path, "message": "Expected string"}]
            enum_values = schema.get("enum") or []
            if enum_values and payload not in enum_values:
                return [{"path": path, "message": f"Expected one of {enum_values}"}]
            return []
        if schema_type == "boolean":
            return [] if isinstance(payload, bool) else [{"path": path, "message": "Expected boolean"}]
        if schema_type == "integer":
            return [] if isinstance(payload, int) and not isinstance(payload, bool) else [{"path": path, "message": "Expected integer"}]
        if schema_type == "number":
            return [] if isinstance(payload, (int, float)) and not isinstance(payload, bool) else [{"path": path, "message": "Expected number"}]
        if isinstance(schema_type, list):
            return []
        return []

    def _execute_workflow(self, workflow: Mapping[str, Any], payload: Mapping[str, Any]) -> Dict[str, Any]:
        workflow_id = str(workflow.get("id") or "")
        if workflow_id == "hazardous_project_list_recognition":
            return self._execute_hazardous_project_list_recognition(workflow, payload)
        if workflow_id == "szzg_agent_encapsulation_route":
            return self._execute_szzg_agent_encapsulation_route(workflow, payload)
        raise RuntimeError(f"No executor registered for workflow: {workflow_id}")

    def _execute_hazardous_project_list_recognition(
        self,
        workflow: Mapping[str, Any],
        payload: Mapping[str, Any],
    ) -> Dict[str, Any]:
        file_info = payload.get("file") if isinstance(payload, Mapping) else None
        if not isinstance(file_info, Mapping):
            raise ValueError("Workflow payload missing file input")

        content_ref = str(file_info.get("content_ref") or file_info.get("path") or file_info.get("file_path") or "").strip()
        if not content_ref:
            raise ValueError("Workflow payload missing file.content_ref")

        input_path = Path(content_ref)
        if not input_path.exists():
            input_path = Path(str(content_ref).replace("\\", "/")).expanduser()
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {content_ref}")

        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise RuntimeError("openpyxl is required for spreadsheet workflows") from exc

        workbook = load_workbook(input_path, data_only=True)
        worksheet = workbook.active
        records: List[Dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            number = row[0] if len(row) > 0 else None
            originname = row[1] if len(row) > 1 else None
            expert_review = row[6] if len(row) > 6 else None
            if number is None and originname is None:
                break
            if number is None or originname is None:
                continue
            originname_text = _clean_string(originname)
            records.append(
                {
                    "id": str(number),
                    "originname": originname_text,
                    "name": self._normalize_hazardous_project_name(originname_text),
                    "category": self._categorize_hazardous_project(originname_text),
                    "isExdanger": _clean_string(expert_review) == "是",
                }
            )
        return {
            "code": 200,
            "msg": "文件解析成功",
            "data": records,
        }

    def _execute_szzg_agent_encapsulation_route(
        self,
        workflow: Mapping[str, Any],
        payload: Mapping[str, Any],
    ) -> Dict[str, Any]:
        if not isinstance(payload, Mapping):
            raise ValueError("Workflow payload must be an object")

        user_input = str(payload.get("user_input") or "").strip()
        candidates = self._normalize_route_candidates(payload.get("data"))

        if self._looks_like_collection_route(user_input):
            return {
                "command": "route_square",
                "params": ["ROUTE://square_page?page=collect"],
                "message": "这就为你打开收藏列表页面",
            }

        if not candidates:
            return {
                "command": "clarify_none",
                "params": [],
                "message": "未找到对应智能体，请确认名称后再试哦",
            }

        if len(candidates) == 1:
            candidate = candidates[0]
            agent_name = str(candidate.get("agentName") or candidate.get("name") or "").strip()
            agent_id = str(candidate.get("agentId") or candidate.get("id") or "").strip()
            if not agent_name or not agent_id:
                raise ValueError("Workflow payload candidate missing agentId or agentName")
            return {
                "command": "route_agent",
                "params": [f"ROUTE://agent_detail?id={agent_id}"],
                "message": f"我马上为你打开{agent_name}智能体",
            }

        agent_names = [str(item.get("agentName") or item.get("name") or "").strip() for item in candidates]
        agent_names = [name for name in agent_names if name]
        return {
            "command": "clarify_multi",
            "params": agent_names,
            "message": "为你找到以下匹配的智能体，请告诉我具体要打开哪一个",
        }

    @staticmethod
    def _normalize_route_candidates(raw_candidates: Any) -> List[Dict[str, Any]]:
        if not isinstance(raw_candidates, list):
            return []
        candidates: List[Dict[str, Any]] = []
        for item in raw_candidates:
            if isinstance(item, Mapping):
                normalized = {
                    "agentId": str(item.get("agentId") or item.get("id") or "").strip(),
                    "agentName": str(item.get("agentName") or item.get("name") or "").strip(),
                }
                if normalized["agentId"] or normalized["agentName"]:
                    candidates.append(normalized)
            elif isinstance(item, str):
                text = item.strip()
                if text:
                    candidates.append({"agentId": "", "agentName": text})
        return candidates

    @staticmethod
    def _looks_like_collection_route(user_input: str) -> bool:
        normalized = user_input.strip()
        return "收藏" in normalized or "收藏列表" in normalized

    @staticmethod
    def _normalize_hazardous_project_name(originname: str) -> str:
        if "专项方案" in originname and not originname.endswith("专项施工方案"):
            return originname.replace("专项方案", "")
        suffix_rules = (
            ("施工组织设计方案", True),
            ("专项施工方案", True),
            ("专项方案", False),
            ("施工方案", False),
            ("方案", False),
        )
        for suffix, append_construction in suffix_rules:
            if originname.endswith(suffix):
                core = originname[: -len(suffix)]
                if append_construction and not core.endswith("吊装"):
                    return f"{core}施工"
                return core
        return originname

    @staticmethod
    def _categorize_hazardous_project(originname: str) -> str:
        if "施工组织设计" in originname:
            return "施工方案管理"
        if "临时用电" in originname:
            return "临时用电工程"
        if "基坑" in originname:
            return "基坑工程"
        if "拆除" in originname or "爆破" in originname:
            return "拆除、爆破工程"
        if "钢箱梁" in originname:
            return "其他类别"
        if any(keyword in originname for keyword in ("现浇箱梁", "现浇盖梁", "支架搭设", "钢管支架")):
            return "模板工程和支架体系"
        if any(keyword in originname for keyword in ("吊装", "架桥机", "架设", "安拆", "钢梁及桥面板安装", "小箱梁架设")):
            return "起重吊装及起重机械安拆工程"
        return "其他类别"

    def _iter_manifest_paths(self) -> Iterable[Path]:
        root = self.root_path
        if not root.exists() or not root.is_dir():
            return []
        manifests: List[Path] = []
        for workflow_dir in sorted((item for item in root.iterdir() if item.is_dir()), key=lambda item: item.name):
            if workflow_dir.name.startswith("_"):
                continue
            yaml_path = workflow_dir / "workflow.yaml"
            yml_path = workflow_dir / "workflow.yml"
            if yaml_path.exists():
                manifests.append(yaml_path)
            elif yml_path.exists():
                manifests.append(yml_path)
        return manifests

    def _load_manifest(self, manifest_path: Path) -> _ManifestResult:
        try:
            raw_manifest = _load_yaml_mapping(manifest_path)
            workflow = self._normalize_manifest(manifest_path, raw_manifest)
            return _ManifestResult(workflow=workflow, error=None)
        except ValueError as exc:
            return _ManifestResult(
                workflow=None,
                error={
                    "status": "invalid",
                    "workflow_dir": str(manifest_path.parent),
                    "manifest_path": str(manifest_path),
                    "message": str(exc),
                },
            )

    def _normalize_manifest(self, manifest_path: Path, manifest: Mapping[str, Any]) -> Dict[str, Any]:
        missing = [field for field in ("id", "name", "version", "owner", "status") if not _clean_string(manifest.get(field))]

        workflow_id = _clean_string(manifest.get("id"))
        status = _clean_string(manifest.get("status"))

        if missing:
            raise ValueError(f"Missing required fields: {', '.join(missing)}")

        owner = _normalize_owner(manifest.get("owner"))
        if not owner.get("primary"):
            raise ValueError("Missing required field: owner.primary")

        source = _normalize_source(manifest.get("source"))
        entrypoint = _normalize_entrypoint(manifest.get("entrypoint"))
        inputs = _normalize_mapping(manifest.get("inputs"))
        outputs = _normalize_mapping(manifest.get("outputs"))
        prompts = _normalize_prompts(manifest.get("prompts"), manifest_path.parent)
        dependencies = _normalize_dependencies(manifest.get("dependencies"))
        governance = _normalize_governance(manifest.get("governance"))
        acceptance = _normalize_acceptance(manifest.get("acceptance"), manifest_path.parent)
        metadata = _normalize_mapping(manifest.get("metadata"))

        readiness = _compute_readiness(
            status=status,
            owner=owner,
            prompts=prompts,
            acceptance=acceptance,
            dependencies=dependencies,
        )

        return {
            "id": workflow_id,
            "name": _clean_string(manifest.get("name")),
            "version": _clean_string(manifest.get("version")),
            "status": status,
            "owner": owner,
            "source": source,
            "entrypoint": entrypoint,
            "inputs": inputs,
            "outputs": outputs,
            "prompts": prompts,
            "dependencies": dependencies,
            "governance": governance,
            "acceptance": acceptance,
            "metadata": metadata,
            "capability_id": entrypoint.get("capability_id") or f"coze.workflow.{workflow_id}",
            "workflow_dir": str(manifest_path.parent),
            "manifest_path": str(manifest_path),
            "readiness": readiness,
        }

    def _build_invalid_contract(self, manifest_path: Path, manifest: Mapping[str, Any], reason: str) -> Dict[str, Any]:
        return {
            "id": _clean_string(manifest.get("id")) or "unknown",
            "name": _clean_string(manifest.get("name")) or "unknown",
            "version": _clean_string(manifest.get("version")) or "unknown",
            "status": "invalid",
            "owner": {},
            "source": {},
            "entrypoint": {},
            "inputs": {},
            "outputs": {},
            "prompts": {},
            "dependencies": {},
            "governance": {},
            "acceptance": {},
            "metadata": {},
            "capability_id": None,
            "workflow_dir": str(manifest_path.parent),
            "manifest_path": str(manifest_path),
            "readiness": {
                "status": "invalid",
                "reason": reason,
                "blockers": ["invalid_manifest"],
            },
        }


def _load_yaml_mapping(path: Path) -> Mapping[str, Any]:
    text = path.read_text(encoding="utf-8")
    try:
        import yaml  # type: ignore
    except ModuleNotFoundError:
        loaded = _parse_limited_yaml(text)
    else:
        loaded = yaml.safe_load(text) or {}
    if not isinstance(loaded, Mapping):
        raise ValueError("Manifest root must be a mapping")
    return loaded


def _normalize_owner(value: Any) -> Dict[str, Any]:
    mapping = _normalize_mapping(value)
    return {
        "team": _clean_string(mapping.get("team")),
        "primary": _clean_string(mapping.get("primary")),
        "reviewers": _normalize_string_list(mapping.get("reviewers")),
    }


def _normalize_source(value: Any) -> Dict[str, Any]:
    mapping = _normalize_mapping(value)
    return {
        "platform": _clean_string(mapping.get("platform")),
        "workspace": _clean_string(mapping.get("workspace")),
        "workflow_id": _clean_string(mapping.get("workflow_id")),
        "workflow_name": _clean_string(mapping.get("workflow_name")),
        "workflow_description": _clean_string(mapping.get("workflow_description")),
        "export_ref": _clean_string(mapping.get("export_ref")),
        "migration_notes": _clean_string(mapping.get("migration_notes")),
        "unsupported_nodes": _normalize_string_list(mapping.get("unsupported_nodes")),
    }


def _normalize_entrypoint(value: Any) -> Dict[str, Any]:
    mapping = _normalize_mapping(value)
    return {
        "mode": _clean_string(mapping.get("mode")),
        "adapter": _clean_string(mapping.get("adapter")),
        "handler": _clean_string(mapping.get("handler")),
        "capability_id": _clean_string(mapping.get("capability_id")),
        "default_timeout_ms": _normalize_optional_int(mapping.get("default_timeout_ms")),
    }


def _normalize_prompts(value: Any, workflow_dir: Path) -> Dict[str, Any]:
    mapping = _normalize_mapping(value)
    prompts = {}
    for key in ("system", "task", "user"):
        path_str = _clean_string(mapping.get(key))
        if path_str:
            prompt_path = workflow_dir / path_str
            exists = prompt_path.exists()
            prompts[key] = {
                "path": path_str,
                "exists": exists,
            }
    return prompts


def _normalize_dependencies(value: Any) -> Dict[str, Any]:
    mapping = _normalize_mapping(value)
    return {
        "tools": _normalize_string_list(mapping.get("tools")),
        "mcp_capabilities": _normalize_string_list(mapping.get("mcp_capabilities")),
        "skills": _normalize_string_list(mapping.get("skills")),
        "providers": _normalize_string_list(mapping.get("providers")),
        "knowledge_sources": _normalize_string_list(mapping.get("knowledge_sources")),
        "runtime_capabilities": _normalize_string_list(mapping.get("runtime_capabilities")),
    }


def _normalize_governance(value: Any) -> Dict[str, Any]:
    mapping = _normalize_mapping(value)
    return {
        "permission_level": _clean_string(mapping.get("permission_level")),
        "trace_required": bool(mapping.get("trace_required")),
        "approval_required": bool(mapping.get("approval_required")),
        "data_sensitivity": _clean_string(mapping.get("data_sensitivity")),
        "allowed_callers": _normalize_string_list(mapping.get("allowed_callers")),
        "policy_notes": _clean_string(mapping.get("policy_notes")),
    }


def _normalize_acceptance(value: Any, workflow_dir: Path) -> Dict[str, Any]:
    mapping = _normalize_mapping(value)
    examples = []
    raw_examples = mapping.get("examples") or []
    if isinstance(raw_examples, list):
        for item in raw_examples:
            if not isinstance(item, Mapping):
                continue
            ex_id = _clean_string(item.get("id"))
            path_str = _clean_string(item.get("path"))
            expected_path_str = _clean_string(item.get("expected_path"))
            required = bool(item.get("required", True))

            path_exists = False
            expected_exists = False
            if path_str:
                path_exists = (workflow_dir / path_str).exists()
            if expected_path_str:
                expected_exists = (workflow_dir / expected_path_str).exists()

            examples.append({
                "id": ex_id,
                "path": path_str,
                "path_exists": path_exists,
                "expected_path": expected_path_str,
                "expected_exists": expected_exists,
                "required": required,
            })

    smoke = _normalize_mapping(mapping.get("smoke"))
    return {
        "examples": examples,
        "smoke": {
            "deterministic": bool(smoke.get("deterministic")),
            "live_model_required": bool(smoke.get("live_model_required")),
            "fixture_row_count": _normalize_optional_int(smoke.get("fixture_row_count")),
        },
    }


def _compute_readiness(
    *,
    status: str,
    owner: Mapping[str, Any],
    prompts: Mapping[str, Any],
    acceptance: Mapping[str, Any],
    dependencies: Mapping[str, Any],
) -> Dict[str, Any]:
    blockers: List[str] = []
    dependencies = dict(dependencies or {})

    if status not in SUPPORTED_STATUSES:
        blockers.append("invalid_status")

    if not owner.get("primary"):
        blockers.append("missing_owner_primary")

    for key, prompt_info in prompts.items():
        if isinstance(prompt_info, dict) and not prompt_info.get("exists"):
            blockers.append(f"missing_prompt_{key}")

    examples = acceptance.get("examples") or []
    missing_required_examples = sum(1 for ex in examples if ex.get("required") and not (ex.get("path_exists") and ex.get("expected_exists")))
    if missing_required_examples > 0:
        blockers.append("missing_required_acceptance_examples")

    runtime_capabilities = [
        str(item or "").strip()
        for item in (dependencies.get("runtime_capabilities") or [])
        if str(item or "").strip()
    ]
    missing_runtime_capabilities = sorted(
        cap for cap in runtime_capabilities if cap not in SUPPORTED_RUNTIME_CAPABILITIES
    )
    if missing_runtime_capabilities:
        blockers.append("missing_runtime_capabilities:" + ",".join(missing_runtime_capabilities))

    if status == "active" and not examples:
        blockers.append("active_status_missing_examples")

    if blockers:
        return {
            "status": "blocked",
            "reason": "Missing required assets or configuration",
            "blockers": sorted(set(blockers)),
        }

    if status == "draft":
        return {
            "status": "draft",
            "reason": "Workflow is in draft status",
            "blockers": [],
        }

    if status == "review":
        return {
            "status": "review",
            "reason": "Workflow is under review",
            "blockers": [],
        }

    if status == "active":
        return {
            "status": "ready",
            "reason": "All required assets present and status is active",
            "blockers": [],
        }

    if status == "deprecated":
        return {
            "status": "deprecated",
            "reason": "Workflow is deprecated",
            "blockers": [],
        }

    if status == "archived":
        return {
            "status": "archived",
            "reason": "Workflow is archived",
            "blockers": [],
        }

    return {
        "status": "unknown",
        "reason": "Unknown workflow status",
        "blockers": [],
    }


def _normalize_mapping(value: Any) -> Dict[str, Any]:
    return dict(value) if isinstance(value, Mapping) else {}


def _normalize_string_list(value: Any) -> List[str]:
    if not isinstance(value, list):
        return []
    return [_clean_string(item) for item in value if _clean_string(item)]


def _normalize_optional_int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _clean_string(value: Any) -> str:
    return str(value or "").strip()


def _parse_limited_yaml(text: str) -> Dict[str, Any]:
    lines = [
        (len(line) - len(line.lstrip(" ")), line.strip())
        for line in text.splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    ]
    if not lines:
        return {}
    parsed, index = _parse_yaml_block(lines, 0, lines[0][0])
    if index != len(lines) or not isinstance(parsed, dict):
        raise ValueError("Unsupported YAML manifest structure")
    return parsed


def _parse_yaml_block(lines: List[tuple[int, str]], index: int, indent: int) -> tuple[Any, int]:
    if index >= len(lines):
        return {}, index
    if lines[index][0] < indent:
        return {}, index
    if lines[index][1].startswith("- "):
        return _parse_yaml_list(lines, index, indent)
    return _parse_yaml_mapping(lines, index, indent)


def _parse_yaml_mapping(lines: List[tuple[int, str]], index: int, indent: int) -> tuple[Dict[str, Any], int]:
    result: Dict[str, Any] = {}
    while index < len(lines):
        current_indent, content = lines[index]
        if current_indent < indent:
            break
        if current_indent > indent:
            raise ValueError(f"Unexpected indentation near: {content}")
        if content.startswith("- "):
            break
        key, value = _split_yaml_key_value(content)
        index += 1
        if value == "":
            if index < len(lines) and lines[index][0] > current_indent:
                child, index = _parse_yaml_block(lines, index, lines[index][0])
                result[key] = child
            else:
                result[key] = None
        else:
            result[key] = _parse_yaml_scalar(value)
    return result, index


def _parse_yaml_list(lines: List[tuple[int, str]], index: int, indent: int) -> tuple[List[Any], int]:
    result: List[Any] = []
    while index < len(lines):
        current_indent, content = lines[index]
        if current_indent < indent:
            break
        if current_indent != indent or not content.startswith("- "):
            break
        item_content = content[2:].strip()
        index += 1
        item: Any
        if item_content == "":
            if index < len(lines) and lines[index][0] > current_indent:
                item, index = _parse_yaml_block(lines, index, lines[index][0])
            else:
                item = None
        elif ":" in item_content:
            key, value = _split_yaml_key_value(item_content)
            item = {key: _parse_yaml_scalar(value)}
            if index < len(lines) and lines[index][0] > current_indent:
                child, index = _parse_yaml_mapping(lines, index, lines[index][0])
                if isinstance(child, dict):
                    item.update(child)
        else:
            item = _parse_yaml_scalar(item_content)
        result.append(item)
    return result, index


def _split_yaml_key_value(content: str) -> tuple[str, str]:
    if ":" not in content:
        raise ValueError(f"Expected key/value pair near: {content}")
    key, value = content.split(":", 1)
    key = key.strip()
    if not key:
        raise ValueError(f"Missing key near: {content}")
    return key, value.strip()


def _parse_yaml_scalar(value: str) -> Any:
    if value in {"", "null", "Null", "NULL", "~"}:
        return None
    if value in {"true", "True", "TRUE"}:
        return True
    if value in {"false", "False", "FALSE"}:
        return False
    if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
        return value[1:-1]
    return value


_coze_workflow_registry_service: Optional[CozeWorkflowRegistryService] = None


def get_coze_workflow_registry_service() -> CozeWorkflowRegistryService:
    global _coze_workflow_registry_service
    if _coze_workflow_registry_service is None:
        _coze_workflow_registry_service = CozeWorkflowRegistryService()
    return _coze_workflow_registry_service
